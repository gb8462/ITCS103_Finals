import customtkinter
import openpyxl
import hashlib, os, re 
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk, ImageEnhance

main = customtkinter.CTk()
main.title("TryQuizMe login")
main.geometry("1100x680")
customtkinter.set_appearance_mode('dark')
customtkinter.set_default_color_theme("blue")

main.configure(fg_color="#1f2024")

quiz_file = "quizzes.xlsx"
if not os.path.exists(quiz_file):
    wb = Workbook()
    sheet = wb.active
    wb.save(quiz_file)

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def valid_username(username):
    return re.match("^[A-Za-z0-9_]{4,}$", username) is not None

def strong_password(password):
    if len(password) < 6:
        return False
    if not re.search(r"[A-Z]", password):
        return False
    if not re.search(r"[a-z]", password):
        return False
    if not re.search(r"[0-9]", password):
        return False
    if not re.search(r"[@$!%*#?&]", password):
        return False
    return True

# ========== openpyxl ==========
def initialize_database():
    if not os.path.exists("users.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        sheet.append(["Username", "Password"])
        workbook.save("users.xlsx")

def save_account(username, password):
    hashed_password = hash_password(password)
    workbook = openpyxl.load_workbook("users.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return False

    sheet.append([username, hashed_password])
    workbook.save("users.xlsx")
    return True

def check_login(username, password):
    hashed_password = hash_password(password)
    workbook = openpyxl.load_workbook("users.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == hashed_password:
            return True
    return False

# ========== Win5w Close ==========
def close_window():
    main.destroy()

# ========== Clear Frame Function ==========
def clear_frame():
    """Clear all widgets in the main window."""
    for widget in main.winfo_children():
        widget.destroy()

# ========== Global Current User ==========
current_user = ""

# ======== Login Page ========
def login_page():
    clear_frame()

    login_frame = customtkinter.CTkFrame(main, corner_radius=20, fg_color="#f7f9fc")
    login_frame.pack(pady=100, padx=300, fill="both", expand=True)

    customtkinter.CTkLabel(
        login_frame, text="Login", font=("Arial", 50, "bold"), text_color="#1f2024"
    ).pack(pady=(40, 20))

    # Username Field
    customtkinter.CTkLabel(login_frame, text="Username", font=("Arial", 13), text_color="#1f2024").pack()
    username_entry = customtkinter.CTkEntry(
        login_frame, width=280, corner_radius=12, border_color="#dcdfe3", text_color="#000000",border_width=2, fg_color="#ffffff"
    )
    username_entry.pack(pady=(5, 10))

    # Password Field
    customtkinter.CTkLabel(login_frame, text="Password", font=("Arial", 13), text_color="#1f2024").pack()
    password_var = customtkinter.StringVar()
    password_entry = customtkinter.CTkEntry(
        login_frame, textvariable=password_var, width=280, corner_radius=12,
        border_color="#dcdfe3", border_width=2, text_color="#000000",fg_color="#ffffff", show="*"
    )
    password_entry.pack(pady=(5, 5))

    # Show/Hide toggle
    def toggle_password_visibility():
        if password_entry.cget("show") == "*":
            password_entry.configure(show="")
            toggle_button.configure(text="Hide")
        else:
            password_entry.configure(show="*")
            toggle_button.configure(text="Show")

    toggle_button = customtkinter.CTkButton(
        login_frame, text="Show", width=60, height=20, font=("Arial", 11), corner_radius=10,
        fg_color="#dcdfe3", text_color="#1f2024", command=toggle_password_visibility
    )
    toggle_button.pack(pady=(0, 15))

    error_label = customtkinter.CTkLabel(login_frame, text="", text_color="#e74c3c", font=("Arial", 12))
    error_label.pack()

    def on_login():
        global current_user
        username = username_entry.get()
        password = password_var.get()

        if check_login(username, password):
            current_user = username
            Dashboard()
        else:
            error_label.configure(text="Invalid username or password")

    main.bind('<Return>', lambda event: on_login())

    # Login Button
    customtkinter.CTkButton(
        login_frame, text="Login", font=("Arial", 16, "bold"), corner_radius=20, text_color="#ffffff",
        fg_color="#4668f2", hover_color="#314ad1", width=150, command=on_login
    ).pack(pady=(20, 0))

    customtkinter.CTkLabel(
        login_frame, text="or", font=("Arial", 12), text_color="#888888"
    ).pack(pady=(5, 5))

    # Sign Up Button
    customtkinter.CTkButton(
        login_frame, text="Sign up", font=("Arial", 15), corner_radius=20, text_color="#4668f2",
        fg_color="#e4e6ed", hover_color="#c9ccd4", border_width=1, border_color="#4668f2", width=150,
        command=signUp_page
    ).pack(pady=(0, 20))


# ======== Sign Up Page ========
def signUp_page():
    clear_frame()

    signup_frame = customtkinter.CTkFrame(main, corner_radius=20, fg_color="#f7f9fc")
    signup_frame.pack(pady=100, padx=300, fill="both", expand=True)

    customtkinter.CTkLabel(
        signup_frame, text="Sign Up", font=("Arial", 50, "bold"), text_color="#1f2024"
    ).pack(pady=(40, 20))

    # Username
    customtkinter.CTkLabel(signup_frame, text="Username", font=("Arial", 13), text_color="#1f2024").pack()
    username_entry = customtkinter.CTkEntry(
        signup_frame, width=280, corner_radius=12, border_color="#dcdfe3",
        border_width=2, fg_color="#ffffff",text_color="#000000"
    )
    username_entry.pack(pady=(5, 10))

    # Password
    customtkinter.CTkLabel(signup_frame, text="Password", font=("Arial", 13), text_color="#1f2024").pack()
    password_var = customtkinter.StringVar()
    password_entry = customtkinter.CTkEntry(
        signup_frame, textvariable=password_var, width=280, corner_radius=12,
        border_color="#dcdfe3", border_width=2, text_color="#000000",fg_color="#ffffff", show="*"
    )
    password_entry.pack(pady=(5, 10))

    # Confirm Password
    customtkinter.CTkLabel(signup_frame, text="Confirm Password", font=("Arial", 13), text_color="#1f2024").pack()
    confirm_password_entry = customtkinter.CTkEntry(
        signup_frame, width=280, corner_radius=12,
        border_color="#dcdfe3", border_width=2, fg_color="#ffffff", text_color="#000000",show="*"
    )
    confirm_password_entry.pack(pady=(5, 5))

    # Toggle Passwords
    def toggle_signup_password():
        if password_entry.cget("show") == "*":
            password_entry.configure(show="")
            confirm_password_entry.configure(show="")
            toggle_signup_button.configure(text="Hide")
        else:
            password_entry.configure(show="*")
            confirm_password_entry.configure(show="*")
            toggle_signup_button.configure(text="Show")

    toggle_signup_button = customtkinter.CTkButton(
        signup_frame, text="Show", width=60, height=20, font=("Arial", 11),
        corner_radius=10, fg_color="#dcdfe3", text_color="#1f2024", command=toggle_signup_password
    )
    toggle_signup_button.pack(pady=(0, 15))

    error_label = customtkinter.CTkLabel(signup_frame, text="", text_color="#e74c3c", font=("Arial", 12))
    error_label.pack()

    def on_create_account():
        username = username_entry.get()
        password = password_var.get()
        confirm_password = confirm_password_entry.get()

        if password != confirm_password:
            error_label.configure(text="Passwords do not match!")
            return

        if save_account(username, password):
            messagebox.showinfo("Success", "Account created successfully!")
            login_page()
        else:
            error_label.configure(text="Username already exists!")

    # Create Account Button
    customtkinter.CTkButton(
        signup_frame, text="Create Account", font=("Arial", 15, "bold"), corner_radius=20, text_color="#ffffff",
        fg_color="#4668f2", hover_color="#314ad1", width=170, command=on_create_account
    ).pack(pady=(10, 5))

    # Back to Login Button
    customtkinter.CTkButton(
        signup_frame, text="Back to Login", font=("Arial", 14), corner_radius=20, text_color="#4668f2",
        fg_color="#e4e6ed", hover_color="#c9ccd4", border_width=1, border_color="#4668f2", width=170,
        command=login_page
    ).pack(pady=(0, 20))

# ========== User's Achievement ==========
achievements = [
    {"name": "Who am i?", "desc": "First badge to achieve after registering an account for the first time", "unlocked": True},
    {"name": "Bullseye", "desc": "100% Accuracy (Get a perfect Score on a Quiz)", "unlocked": False},
    {"name": "The Fool", "desc" : "Scored 0 on a Quiz", "unlocked":True},
    {"name": "Now that's what I call  a Purrrfectionist!", "desc": "Get three perfect scores in three different Categories", "unlocked":False},
    {"name": "IT Specialist", "desc": "Get a perfect score on the General Category in TryQuizMe", "unlocked":False},
    {"name": "Achievement Hoarder", "desc": "Collect 20 Achievements", "unlocked":False},
    {"name": "His/Her Majesty", "desc": "Ranked 1st in the leaderboard", "unlocked":False},
    {"name": "Duke? I thought you meant Duck!", "desc": "Ranked 2nd in the leaderboard", "unlocked":False},
    {"name": "Marquess", "desc": "Ranked 3rd in the leaderboard", "unlocked":False},
    {"name": "Nobles", "desc": "Ranked 4th or 5th in the leaderboard", "unlocked":False},
    {"name": "First Steps", "desc": "5th Milestone", "unlocked": False},
    {"name": "Nerd", "desc": "20th Milestone", "unlocked": False},
    {"name": "Touch Grass", "desc": "50th Milestone", "unlocked": False},
    {"name": "Sparkly shimmery splendid etc!", "desc": "100th Milestone", "unlocked": False},
    {"name": "Moar stars to come!", "desc": "150th Milestone", "unlocked": False},
]
# ========== Achievements ===========

# Create placeholder images
def create_placeholder_image(color):
    img = Image.new("RGB", (64, 64), color=color)
    return img

# Apply dimming to locked achievements
def process_image(img, unlocked):
    if not unlocked:
        enhancer = ImageEnhance.Brightness(img)
        img = enhancer.enhance(0.3)  # Dim the image
    return ImageTk.PhotoImage(img)

def Achievements():
    clear_frame()
    main.title("TryQuizMe")
    main.configure(fg_color="#010101")

    frame = customtkinter.CTkFrame(main, fg_color="#353A3E", corner_radius=0)
    frame.pack(fill='both', expand=True)

    title = customtkinter.CTkLabel(frame, text="🌟 Achievements 🌟", font=("Arial", 30), text_color="#ffffff")
    title.pack(pady=(30, 10))

    tree_frame = customtkinter.CTkScrollableFrame(frame, width=1000, height=400, fg_color="#1e1e2f")
    tree_frame.pack(pady=10, fill="both", expand=True)

    last_frame = customtkinter.CTkFrame(frame)
    last_frame.pack(padx=100,pady=10)

    for ach in achievements:
        frame = ttk.Frame(tree_frame, padding=5)
        frame.pack(fill="x", pady=5)

        # Create and process the image
        base_img = create_placeholder_image("green" if ach["unlocked"] else "gray")
        photo = process_image(base_img, ach["unlocked"])

        # Image on left
        img_label = ttk.Label(frame, image=photo)
        img_label.image = photo  # Keep a reference
        img_label.pack(side="left")

        # Text on right
        text_frame = ttk.Frame(frame)
        text_frame.pack(side="left", padx=10)

        name_style = {'foreground': 'black' if ach['unlocked'] else 'gray'}
        desc_style = {'foreground': 'black' if ach['unlocked'] else 'gray'}

        name_label = ttk.Label(text_frame, text=ach["name"], font=("Arial", 12, "bold"), **name_style)
        desc_label = ttk.Label(text_frame, text=ach["desc"], font=("Arial", 10), **desc_style)

        name_label.pack(anchor="w")
        desc_label.pack(anchor="w")

    customtkinter.CTkButton(last_frame,text="Back to Dashboard", command=Dashboard, fg_color="#4668f2", hover_color="#314ad1", text_color="#fff",font=("Arial", 14), corner_radius=8, width=180).pack(pady=(0, 0))

# ========== Leaderboard ==========
def Leaderboard():
    clear_frame()
    main.title("TryQuizMe")
    main.configure(fg_color="#010101")

    frame = customtkinter.CTkFrame(main, fg_color="#353A3E", corner_radius=0)
    frame.pack(fill='both', expand=True)

    title = customtkinter.CTkLabel(frame, text="🏆 Leaderboard 🏆", font=("Arial", 30), text_color="#ffffff")
    title.pack(pady=(30, 10))

    tree_frame = customtkinter.CTkFrame(frame, fg_color="#2e2f33", corner_radius=15)
    tree_frame.pack(padx=60, pady=20, fill="both", expand=True)

    style = ttk.Style()
    style.theme_use("default")
    style.configure("Treeview", background="#2e2f33", foreground="#f2f2f2", rowheight=30, fieldbackground="#2e2f33", font=("Arial", 12))
    style.configure("Treeview.Heading", font=("Arial", 16, "bold"), background="#353A3E", foreground="#ffffff")
    style.map("Treeview", background=[('selected', '#353A3E')], foreground=[('selected', '#ffffff')])

    tree = ttk.Treeview(tree_frame, columns=("Rank", "Username", "Score"), show="headings", selectmode="none")
    tree.heading("Rank", text="Rank")
    tree.heading("Username", text="Username")
    tree.heading("Score", text="Score")
    tree.column("Rank", anchor="center", width=80)
    tree.column("Username", anchor="w", width=200)
    tree.column("Score", anchor="center", width=100)

    tree.pack(fill="both", expand=True, padx=20, pady=20)

    # Insert data
    try:
        wb = openpyxl.load_workbook("users.xlsx")
        sheet = wb.active

        users_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            username = row[0]
            score = row[2] if len(row) > 2 and isinstance(row[2], int) else 0
            users_data.append((username, score))

        users_data.sort(key=lambda x: x[1], reverse=True)

        for index, (username, score) in enumerate(users_data, start=1):
            tree.insert("", "end", values=(index, username, score))

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load leaderboard: {e}")

    customtkinter.CTkButton(frame, text="Back to Dashboard", command=Dashboard, fg_color="#4668f2", hover_color="#314ad1", text_color="#fff",font=("Arial", 14), corner_radius=8, width=180).pack(pady=(0, 20))

# ========== Dashboard ==========
def Dashboard():
    clear_frame()
    main.title("TryQuizMe")

    topbar = customtkinter.CTkFrame(main, fg_color="#353A3E", corner_radius=0, height=60)
    topbar.pack(pady=0, padx=0, fill="both")

    # Load the logo image
    logo_path = os.path.join("assets", "logo.png")
    logo_img = customtkinter.CTkImage(light_image=Image.open(logo_path), size=(40, 40))
    logo_label = customtkinter.CTkLabel(topbar, image=logo_img, text="")  # no text
    logo_label.pack(pady=10, padx=(20, 0), side="left")

    # App title
    title_label = customtkinter.CTkLabel(topbar, text="TryQuizMe", font=('Arial', 25))
    title_label.pack(pady=10, padx=(5, 0), side='left')

    # Load the profile pic
    pfp_path = os.path.join("assets", "pfp.png")
    pfp_img = customtkinter.CTkImage(light_image=Image.open(pfp_path), size=(40, 40))
    pfp_label = customtkinter.CTkLabel(topbar, image=pfp_img, text="")
    pfp_label.pack(pady=10, padx=(0, 10), side="right")

    # Username label
    user_label = customtkinter.CTkLabel(topbar, text=current_user, font=('Arial', 18))
    user_label.pack(pady=10, padx=(0, 5), side='right')

    # Background
    background = customtkinter.CTkFrame(main, fg_color="#f0f0f0", corner_radius=0, height=400)
    background.pack(fill="both", expand=True)

    # Buttons
    customtkinter.CTkButton(background, text="QuizMe", command=quizMe, corner_radius=3, width=350, height=40, font=('Arial', 20), fg_color='#353A3E', text_color='#E0E0E0', hover_color='#1A1A1A').pack(pady=(170, 5))
    customtkinter.CTkButton(background, text="Achievements", command=Achievements, corner_radius=3, width=350, height=40, font=('Arial', 20), fg_color='#353A3E', text_color='#E0E0E0', hover_color='#1A1A1A').pack(pady=(10, 5))
    customtkinter.CTkButton(background, text="Leaderboards", command=Leaderboard, corner_radius=3, width=350, height=40, font=('Arial', 20), fg_color='#353A3E', text_color='#E0E0E0', hover_color='#1A1A1A').pack(pady=(10, 5))
    customtkinter.CTkButton(background, text="Logout", command=login_page, corner_radius=3, width=350, height=40, font=('Arial', 20), fg_color='#353A3E', text_color='#E0E0E0', hover_color='#1A1A1A').pack(pady=(10, 5))


# Global lists for Create Quiz

isho_quiz_name_entry = None
isho_question_entries = []
isho_choice_entries = []
isho_correct_answers = []

# ========== QuizMe ==========
def quizMe():
    clear_frame()
    main.title("QuizMe")

    # ========== Dashboard (Difficulty/Category Filter) ==========
    dashboard = customtkinter.CTkFrame(main, corner_radius=10, fg_color="#dee0e0", width=300)
    dashboard.pack(pady=(5, 20), padx=10, fill="both", side="left")

    customtkinter.CTkLabel(dashboard, text="QuizMe", font=("Arial", 35), text_color="#101010").pack(pady=(15, 15), padx=70)

    customtkinter.CTkLabel(dashboard, text="Difficulty", font=("Arial", 17), text_color="#101010").pack(pady=(25, 0), padx=(0, 135))
    difficulty = customtkinter.CTkFrame(dashboard, fg_color="#ebeded", border_width=1, border_color="#c9c9c9")
    difficulty.pack()

    customtkinter.CTkLabel(dashboard, text="Category", font=("Arial", 17), text_color="#101010").pack(pady=(25, 0), padx=(0, 135))
    category = customtkinter.CTkFrame(dashboard, fg_color="#ebeded", border_width=1, border_color="#c9c9c9")
    category.pack(pady=0, padx=1)

    # Track selected filter state
    selected_difficulty = customtkinter.StringVar(value="All Difficulty")
    selected_category = customtkinter.StringVar(value="All Category")

    # Dictionaries to hold button references
    difficulty_buttons = {}
    category_buttons = {}

    # Function to update active button styles
    def update_filter_buttons():
        for d, btn in difficulty_buttons.items():
            if d == selected_difficulty.get():
                btn.configure(fg_color="#4668f2", text_color="#ffffff")
            else:
                btn.configure(fg_color="#ffffff", text_color="#101010")

        for c, btn in category_buttons.items():
            if c == selected_category.get():
                btn.configure(fg_color="#4668f2", text_color="#ffffff")
            else:
                btn.configure(fg_color="#ffffff", text_color="#101010")

    # Filter handlers
    def select_difficulty(d):
        selected_difficulty.set(d)
        update_filter_buttons()
        reload_quiz_list()

    def select_category(c):
        selected_category.set(c)
        update_filter_buttons()
        reload_quiz_list()

    # Create difficulty buttons
    for diff in ["All Difficulty", "Easy", "Medium", "Hard"]:
        btn = customtkinter.CTkButton( difficulty, text=diff, anchor="w", height=35, corner_radius=0, width=200, fg_color="#ffffff", text_color="#101010", hover_color="#4668f2", command=lambda d=diff: select_difficulty(d))
        btn.pack(pady=1, padx=1)
        difficulty_buttons[diff] = btn

    # Create category buttons
    for cat in ["All Category", "General", "Web Development", "Cryptography", "Python"]:
        btn = customtkinter.CTkButton( category, text=cat, anchor="w", height=35, corner_radius=0, width=200, fg_color="#ffffff", text_color="#101010", hover_color="#4668f2", command=lambda c=cat: select_category(c))
        btn.pack(pady=1, padx=1)
        category_buttons[cat] = btn

    # Highlight the initial selected filters
    update_filter_buttons()

    # ========== Create Quiz ==========
    def isho_create_quiz_page(quiz_name, existing_data, metadata):
        clear_frame()
        main.configure(fg_color="#1e1e2f")

        global isho_quiz_name_entry, isho_question_entries, isho_choice_entries, isho_correct_answers
        isho_question_entries, isho_choice_entries, isho_correct_answers = [], [], []

        metadata = existing_data[0] if existing_data else None
        quiz_data = existing_data[1:] if existing_data else None

        customtkinter.CTkLabel(main, text="Create Quiz", font=("Segoe UI", 24), text_color="#ffffff").pack(pady=10)

        # Quiz Name
        customtkinter.CTkLabel(main, text="Quiz Name:", text_color="#ffffff", font=("Segoe UI", 14)).pack(pady=5)
        isho_quiz_name_entry = customtkinter.CTkEntry(main, width=300, height=30, font=("Segoe UI", 13), fg_color="#33334d", text_color="#ffffff")
        if quiz_name:
            isho_quiz_name_entry.insert(0, quiz_name)
            isho_quiz_name_entry.configure(state="disabled")
        isho_quiz_name_entry.pack(pady=5)

        # Dropdowns
        dropdown_row = customtkinter.CTkFrame(main, fg_color="transparent")
        dropdown_row.pack(pady=5)

        customtkinter.CTkLabel(dropdown_row, text="Difficulty:", text_color="#ffffff", font=("Segoe UI", 13)).pack(side="left", padx=(0, 5))
        isho_difficulty = customtkinter.CTkOptionMenu( dropdown_row, values=["Easy", "Medium", "Hard"], width=120, font=("Segoe UI", 12), fg_color="#33334d", button_color="#5e60ce", dropdown_fg_color="#29293d", dropdown_hover_color="#4a4ccc", text_color="#ffffff", dropdown_text_color="#ffffff")
        isho_difficulty.pack(side="left", padx=(0, 20))

        customtkinter.CTkLabel(dropdown_row, text="Category:", text_color="#ffffff", font=("Segoe UI", 13)).pack(side="left", padx=(0, 5))
        isho_category = customtkinter.CTkOptionMenu( dropdown_row, values=["General", "Web Development", "Cryptography", "Python"], width=160, font=("Segoe UI", 12), fg_color="#33334d", button_color="#5e60ce", dropdown_fg_color="#29293d", dropdown_hover_color="#4a4ccc", text_color="#ffffff", dropdown_text_color="#ffffff")
        isho_category.pack(side="left")

        # Defaults
        isho_difficulty.set(metadata[1] if metadata else "Easy")
        isho_category.set(metadata[2] if metadata else "General")

        questions_container = customtkinter.CTkScrollableFrame(main, width=1000, height=400, fg_color="#1e1e2f")
        questions_container.pack(pady=10, fill="both", expand=True)

        def add_question(prefill=None):
            frame = customtkinter.CTkFrame(questions_container, fg_color="#29293d")
            frame.pack(pady=15, fill="x", padx=20)

            q_entry = customtkinter.CTkEntry(frame, width=600, placeholder_text="Enter question", font=("Segoe UI", 12), fg_color="#33334d", text_color="#ffffff")
            q_entry.pack(pady=6)
            if prefill: q_entry.insert(0, prefill[0])

            choices = []
            for i in range(4):
                entry = customtkinter.CTkEntry(frame, width=500, placeholder_text=f"Choice {chr(65 + i)}", font=("Segoe UI", 12), fg_color="#33334d", text_color="#ffffff")
                entry.pack(pady=5)
                if prefill: entry.insert(0, prefill[i + 1])
                choices.append(entry)

            # Horizontal correct answer row
            correct_var = customtkinter.IntVar(value=prefill[5] if prefill else 0)
            correct_frame = customtkinter.CTkFrame(frame, fg_color="transparent")
            correct_frame.pack(pady=10)

            customtkinter.CTkLabel(correct_frame, text="Correct Choice:", text_color="#cccccc", font=("Segoe UI", 12)).pack(pady=5)

            radio_row = customtkinter.CTkFrame(correct_frame, fg_color="transparent")
            radio_row.pack()

            for i in range(4):
                customtkinter.CTkRadioButton( radio_row, text=f"{chr(65 + i)}", variable=correct_var, value=i, text_color="#cccccc", fg_color="#5e60ce", hover_color="#4a4ccc").pack(side="left", padx=8)

            isho_question_entries.append(q_entry)
            isho_choice_entries.append(choices)
            isho_correct_answers.append(correct_var)

        def remove_last_question():
            if isho_question_entries:
                isho_question_entries.pop().master.destroy()
                isho_choice_entries.pop()
                isho_correct_answers.pop()

        def save_quiz():
            name = isho_quiz_name_entry.get().strip()
            if not name:
                messagebox.showerror("Error", "Please enter a quiz name.")
                return

            wb = load_workbook(quiz_file)

            if  name in wb.sheetnames and not existing_data:
                messagebox.showerror("Error", "Quiz already exists.")
                return

            if name in wb.sheetnames:
                del wb[name]

            sheet = wb.create_sheet(title=name)
            sheet.append([name, isho_difficulty.get(), isho_category.get()])

            for q_entry, choices, correct in zip(isho_question_entries, isho_choice_entries, isho_correct_answers):
                question = q_entry.get().strip()
                choice_vals = [c.get().strip() for c in choices]

                if not question or any(c == "" for c in choice_vals):
                    messagebox.showerror("Error", "All questions and choices must be filled.")
                    return

                sheet.append([question] + choice_vals + [correct.get()])

            wb.save(quiz_file)
            messagebox.showinfo("Success", "Quiz saved successfully!")
            quizMe()

        # Load existing questions
        if quiz_data:
            for q in quiz_data:
                add_question(q)
        else:
            add_question()

        # Add/Remove/Save buttons
        controls = customtkinter.CTkFrame(main, fg_color="transparent")
        controls.pack(pady=5)

        customtkinter.CTkButton(controls, text="+ Add Question", command=add_question, fg_color="#38b000", hover_color="#2e8b00", font=("Segoe UI", 12)).pack(side="left", padx=10)
        customtkinter.CTkButton(controls, text="- Remove Last", command=remove_last_question, fg_color="#e63946", hover_color="#c82333", font=("Segoe UI", 12)).pack(side="left", padx=10)
        customtkinter.CTkButton(controls, text="💾 Save Quiz", command=save_quiz, fg_color="#5e60ce", hover_color="#4a4ccc", font=("Segoe UI", 12)).pack(side="left", padx=10)
        customtkinter.CTkButton(main, text="Go Back", command=quizMe, fg_color="#6c6c6c", hover_color="#4d4d4d", font=("Segoe UI", 12), corner_radius=10, width=150).pack(pady=3)

    def open_edit_quiz_page():
        clear_frame()
        main.configure(fg_color="#1f2024")

        customtkinter.CTkLabel(main, text="Edit Quiz", font=("Arial", 28), text_color="#ffffff").pack(pady=(20, 10))

        try:
            wb = load_workbook(quiz_file)
            quiz_names = wb.sheetnames
        except FileNotFoundError:
            quiz_names = []

        # Frame for quiz selection
        selection_frame = customtkinter.CTkFrame(main, fg_color="#2a2b2e", corner_radius=12)
        selection_frame.pack(pady=10, padx=20, fill="both", expand=True)

        quiz_listbox = customtkinter.CTkScrollableFrame(selection_frame, width=550, height=400, fg_color="#3a3b3e")
        quiz_listbox.pack(pady=20,padx=20)

        customtkinter.CTkLabel(quiz_listbox, text="Select A Quiz To Edit", font=("Arial", 16), text_color="#ffffff").pack(pady=(10, 5))
        selected_quiz = customtkinter.StringVar(value="")

        def select_quiz(name):
            selected_quiz.set(name)

            try:
                wb = load_workbook(quiz_file)
                if name in wb.sheetnames:
                    sheet = wb[name]
                    rows = list(sheet.iter_rows(values_only=True))
                    if rows:
                        metadata = rows[0]  # [quiz name, difficulty, category]
                        questions_data = rows[1:]  # remaining rows are the questions
                        isho_create_quiz_page(name, [metadata] + questions_data, metadata)
                    else:
                        messagebox.showerror("Error", "Quiz data is empty.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load quiz: {e}")

        for q in quiz_names:
            btn = customtkinter.CTkButton(quiz_listbox, text=q, anchor="w", fg_color="#444", text_color="#ffffff", hover_color="#4668f2", width=460, command=lambda name=q: select_quiz(name))
            btn.pack(pady=3, padx=10)
        
        def delQ():
            clear_frame()
            customtkinter.CTkLabel(main, text="Delete Quiz", font=("Arial", 28), text_color="#ffffff").pack(pady=(20, 10))
            # Reload workbook and get quiz names
            try:
                wb = load_workbook(quiz_file)
                quiz_names = wb.sheetnames
            except FileNotFoundError:
                quiz_names = []

            selection_frame = customtkinter.CTkFrame(main, fg_color="#2a2b2e", corner_radius=12)
            selection_frame.pack(pady=10, padx=20, fill="both", expand=True)

            # Scrollable list of quizzes
            quiz_list_frame = customtkinter.CTkScrollableFrame(selection_frame, width=550, height=400, fg_color="#3a3b3e")
            quiz_list_frame.pack(pady=(20, 10), padx=20)

            customtkinter.CTkLabel(quiz_list_frame, text="Select a Quiz to Delete", font=("Arial", 16), text_color="#ffffff").pack(pady=(10, 5))

            quiz_entry = customtkinter.CTkEntry(selection_frame, width=300, corner_radius=8, font=("Arial", 13), placeholder_text="e.g. Python 1")
            quiz_entry.pack(pady=(0, 12))

            # Quiz button click fills the entry
            for q in quiz_names:
                customtkinter.CTkButton(quiz_list_frame,text=q,anchor="w",fg_color="#444",text_color="#ffffff",hover_color="#dc2626",width=460,command=lambda name=q: quiz_entry.delete(0, 'end') or quiz_entry.insert(0, name)).pack(pady=3, padx=10)

            def delete_quiz():
                name = quiz_entry.get().strip()
                if not name:
                    messagebox.showerror("Error", "Please enter a quiz name to delete.")
                    return
                if name not in wb.sheetnames:
                    messagebox.showerror("Error", f"Quiz '{name}' does not exist.")
                    return
                confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete the quiz '{name}'?")
                if confirm:
                    del wb[name]
                    wb.save(quiz_file)
                    messagebox.showinfo("Deleted", f"Quiz '{name}' was deleted successfully!")
                    quizMe()

            # Buttons
            button_row = customtkinter.CTkFrame(selection_frame, fg_color="transparent")
            button_row.pack(pady=(10, 10))

            customtkinter.CTkButton(button_row, text="🗑 Delete", command=delete_quiz, fg_color="#ef4444", hover_color="#dc2626", text_color="#ffffff", width=140, corner_radius=8, font=("Arial", 14, "bold")).pack(side="left", padx=6)
            customtkinter.CTkButton(button_row, text="← Go Back", command=open_edit_quiz_page, fg_color="#9ca3af", hover_color="#6b7280", text_color="#000000", width=140, corner_radius=8, font=("Arial", 14)).pack(side="left", padx=6)

        # Buttons
        button_row = customtkinter.CTkFrame(selection_frame, fg_color="transparent")
        button_row.pack(pady=(10, 10))

        customtkinter.CTkButton(button_row, text="🗑 Delete A Quiz", command=delQ, fg_color="#ef4444", hover_color="#dc2626", text_color="#ffffff", width=140, corner_radius=8, font=("Arial", 14, "bold")).pack(side="left", padx=6)
        customtkinter.CTkButton(button_row, text="← Go Back", command=quizMe, fg_color="#9ca3af", hover_color="#6b7280", text_color="#000000", width=140, corner_radius=8, font=("Arial", 14)).pack(side="left", padx=6)

    # ========== Buttons ==========
    CreateQ = customtkinter.CTkButton(dashboard, text='Create Quiz', command=lambda: isho_create_quiz_page(quiz_name="", existing_data=[], metadata=["Category", "Difficulty"]))
    CreateQ.pack(pady=(15,0))
    
    customtkinter.CTkButton(dashboard, text="Edit Quiz", command=open_edit_quiz_page).pack(pady=(15,0))

    back = customtkinter.CTkButton(dashboard, text='Go Back', command=Dashboard)
    back.pack(pady=(15,0))

    # ========== Take Quiz Available ==========
    Quiz = customtkinter.CTkFrame(main, corner_radius=10, fg_color='#dee0e0')
    Quiz.pack(pady=(5, 20), padx=(0, 10), side='right', fill="both", expand=True)

    label = customtkinter.CTkLabel( Quiz, text="Available Quizzes", font=("Arial", 24, "bold"), text_color="#1f2024")
    label.pack(pady=(20, 10))

    # Create an inner frame for quiz buttons using grid layout
    quiz_grid_frame = customtkinter.CTkFrame(Quiz, fg_color="transparent")
    quiz_grid_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # Make 3 columns responsive
    for i in range(3):
        quiz_grid_frame.grid_columnconfigure(i, weight=1)

    def reload_quiz_list():
        # Clear previous quiz buttons
        for widget in quiz_grid_frame.winfo_children():
            widget.destroy()

        try:
            wb = load_workbook(quiz_file)
            excluded_sheets = ["template", "scores", "users", "sheet", "filters"]

            row = 0
            col = 0

            for sheet_name in wb.sheetnames:
                if sheet_name.lower() in excluded_sheets:
                    continue

                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                if not rows or len(rows[0]) < 3:
                    continue

                quiz_name = rows[0][0] or sheet_name
                quiz_diff = rows[0][1]
                quiz_cat = rows[0][2]

                if selected_difficulty.get() not in ("", "All Difficulty") and quiz_diff != selected_difficulty.get():
                    continue
                if selected_category.get() not in ("", "All Category") and quiz_cat != selected_category.get():
                    continue

                quiz_button = customtkinter.CTkButton(quiz_grid_frame,text=quiz_name, command=lambda name=quiz_name: take_quiz(name), fg_color="#4668f2", hover_color="#314ad1", text_color="#ffffff", font=("Arial", 15), width=200, height=45, corner_radius=12)
                quiz_button.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

                col += 1
                if col > 2:
                    col = 0
                    row += 1

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load quizzes: {str(e)}")  

    def take_quiz(quiz_name):
        clear_frame()
        main.configure(fg_color="#1f2024")

        # Quiz title
        customtkinter.CTkLabel(main, text=f"{quiz_name} Quiz", font=("Arial", 28), text_color="#ffffff").pack(pady=20)

        # Scrollable container for questions
        quiz_frame = customtkinter.CTkScrollableFrame(main, width=800, height=500, fg_color="#2a2b2e")
        quiz_frame.pack(pady=10)

        wb = load_workbook(quiz_file)
        sheet = wb[quiz_name]

        selected_answers, valid_rows = [], []

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            if not row or len(row) < 6:
                print(f"⚠️ Skipping row {i} due to wrong length: {row}")
                continue

            q_text, choice_a, choice_b, choice_c, choice_d, correct_index = row[:6]
            choices = [choice_a, choice_b, choice_c, choice_d]
            valid_rows.append((i, correct_index))

            # Question text
            customtkinter.CTkLabel(quiz_frame, text=f"Q{i}: {q_text}", font=("Arial", 18), text_color="#ffffff", wraplength=700, justify="left").pack(anchor="w", pady=(15, 5), padx=20)

            var = customtkinter.IntVar(value=-1)
            selected_answers.append(var)

            # Answer choices
            for idx, choice in enumerate(choices):
                if choice is None:
                    continue
                customtkinter.CTkRadioButton(quiz_frame, text=choice, variable=var, value=idx, text_color="#ccc", hover_color="#4668f2", fg_color="#76b5c5", font=("Arial", 14)).pack(anchor="w", padx=40, pady=2)

            ttk.Separator(quiz_frame, orient="horizontal").pack(fill="x", pady=10, padx=20)

        # Submission logic
        def submit_answers():
            score = sum(1 for i, (_, correct) in enumerate(valid_rows) if selected_answers[i].get() == correct)
            messagebox.showinfo("Quiz Result", f"You got {score} out of {len(selected_answers)} correct!")

            try:
                user_wb = openpyxl.load_workbook("users.xlsx")
                user_sheet = user_wb.active
                headers = [cell.value for cell in user_sheet[1]]

                score_col_index = headers.index("Score") + 1 if "Score" in headers else len(headers) + 1
                if "Score" not in headers:
                    user_sheet.cell(row=1, column=score_col_index, value="Score")

                for row in user_sheet.iter_rows(min_row=2):
                    if row[0].value == current_user:
                        current_score = row[score_col_index - 1].value or 0
                        user_sheet.cell(row=row[0].row, column=score_col_index, value=current_score + score)
                        break

                user_wb.save("users.xlsx")
                quizMe()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to save score: {e}")

        # Action buttons
        customtkinter.CTkButton(main, text="Submit Quiz", command=submit_answers, fg_color="#4668f2", hover_color="#314ad1", font=("Arial", 12), corner_radius=10, width=200).pack(pady=5)
        customtkinter.CTkButton(main, text="Go Back", command=quizMe, fg_color="#6c6c6c", hover_color="#3a3a3a", font=("Arial", 12), corner_radius=10, width=150).pack(pady=5)

    # Load quizzes on start
    reload_quiz_list()

# ========== Start ==========
initialize_database()
login_page()

main.mainloop()