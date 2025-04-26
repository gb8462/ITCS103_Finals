import os
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from Modules import (takeQuiz, Credits, createQuiz)

window = tk.Tk()
window.geometry("900x500")
window.title("TryQuizMe")
window.configure(bg="#353A3E")

# Set up excel file if it doesn't exist
if not os.path.exists('users.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Password"])
    wb.save('users.xlsx')

# Functions
def TQuiz():
    takeQuiz.quiz(window)

def CQuiz():
    createQuiz.makeQuiz(window)

def cred():
    Credits.credits(window)

def close_window():
    window.destroy()

def load_users():
    wb = load_workbook('users.xlsx')
    ws = wb.active
    users = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        users[row[0]] = row[1]
    return users

def save_user(name, password):
    wb = load_workbook('users.xlsx')
    ws = wb.active
    ws.append([name, password])
    wb.save('users.xlsx')

def clear_window():
    for widget in window.winfo_children():
        widget.destroy()

# Login Functions
def login_screen():
    clear_window()

    label = ttk.Label(window, text="TryQuizMe", font=("Roboto", 30), foreground="#EAEAEA", background="#353A3E")
    label.pack(pady=25)
    label = tk.Label(window, text="Login", font=("Arial",20), foreground="#EAEAEA", background="#353A3E")
    label.pack(pady=5)


    tk.Label(window, text="Username:", bg="#353A3E", fg="#EAEAEA").pack()
    global username_entry
    username_entry = tk.Entry(window,width=40)
    username_entry.pack(pady=5)

    tk.Label(window, text="Password:", bg="#353A3E", fg="#EAEAEA").pack()
    global password_entry
    password_entry = tk.Entry(window, width=40,show="*")
    password_entry.pack(pady=5)

    tk.Button(window, text="Sign In", width=20, command=sign_in).pack(pady=15)
    tk.Button(window, text="Sign Up", width=20, command=sign_up).pack(pady=5)

def sign_in():
    name = username_entry.get()
    password = password_entry.get()
    users = load_users()

    if name in users and users[name] == password:
        messagebox.showinfo("Success", f"Welcome back {name}!")
        main()  # <--- move to main menu after login
    else:
        messagebox.showerror("Error", "Invalid username or password!")

def sign_up():
    name = username_entry.get()
    password = password_entry.get()
    users = load_users()

    if name in users:
        messagebox.showerror("Error", "Username already exists!")
    else:
        save_user(name, password)
        messagebox.showinfo("Success", "Account created successfully!")

# Main Window
def main():
    clear_window()

    label = ttk.Label(window, text="TryQuizMe", font=("Roboto", 40), foreground="#EAEAEA", background="#353A3E")
    label.pack(pady=25)

    label = ttk.Label(window, text="Highest Score", font=("Roboto", 12), foreground="#EAEAEA", background="#353A3E")
    label.pack(pady=25)

    button1 = tk.Button(window, text="Take Quiz", height=3, width=25, command=TQuiz)
    button1.pack(pady=10)

    button2 = tk.Button(window, text="Create Quiz", height=3, width=25, command=CQuiz)
    button2.pack(pady=10)

    button3 = tk.Button(window, text="Quit", height=3, width=25, command=close_window)
    button3.pack(pady=10)

# Start with the login screen
login_screen()

window.mainloop()
